import json
from datetime import timedelta
import openpyxl
from django.db.models import Q
from openpyxl.styles import Font, Border, Side, PatternFill, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django_redis import get_redis_connection
from .models import Rental, Car, Shift, BuildObject
from .forms import DocumentForm
import calendar

#TODO: добавить свободный файл в загрузку
#TODO: добавить несколько документов одного типа


def upload_documents(request, pk):
    rental = get_object_or_404(Rental, pk=pk)
    if request.method == 'POST':
        document_form = DocumentForm(request.POST, request.FILES)
        if document_form.is_valid():
            # Сохраняем каждый загружаемый файл отдельно
            for field_name, file in request.FILES.items():
                # Получаем соответствующее поле модели Rental по его имени
                field = getattr(rental, field_name)
                # Сохраняем файл в соответствующее поле модели Rental
                field.save(file.name, file, save=True)

            # После успешной загрузки документов, перенаправляем на страницу с деталями аренды
            return redirect('rental_detail', pk=pk)  # Замените 'rental_detail' на ваш URL pattern

        else:
            # Если форма невалидна, вернем ошибку с соответствующим статусом
            # и передадим параметры ошибки в URL для дальнейшего отображения на странице
            error_params = "?error=invalid_form"
            return redirect('rental_detail', pk=pk) + error_params  # Замените 'rental_detail' на ваш URL pattern

    # Если метод запроса не POST, вернем ошибку с соответствующим статусом
    return JsonResponse({'error': 'Метод не поддерживается'}, status=405)


def generate_excel(request, selected_date):
    start_date = selected_date.replace(day=1)
    end_date = start_date.replace(month=start_date.month % 12 + 1, day=1)
    _, num_days = calendar.monthrange(selected_date.year, selected_date.month)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Special Equipment Report'

    # Set font styles
    header_font = Font(bold=True, size=14)
    menu_font = Font(bold=True, size=14)
    cell_font = Font(size=12)
    total_font = Font(size=18)

    # Set fill styles
    menu_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Set border styles
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    medium_border = Border(left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium"))

    # Add headers with styles
    headers = ['Номер', 'Техника', 'Объект', 'Тариф', 'Доп. тариф']
    for day in range(1, num_days + 1):
        headers.append(day)
    headers += ['количество часов', 'всего (р.)']
    ws.append(headers)

    # Apply font styles and borders to header
    for cell in ws[1]:
        cell.font = header_font
        cell.border = thin_border

    # Получаем данные об арендах и сменах
    rentals = Rental.objects.filter(Q(start_date__lte=end_date) & Q(end_date__gte=start_date))
    shifts = Shift.objects.filter(rental__in=rentals)
    all_cars = Car.objects.all()

    # Получаем список машин, для которых есть подходящие аренды в выбранном месяце
    cars_with_rentals = rentals.values_list('car', flat=True).distinct()

    # Отфильтровываем машины, для которых нет подходящих аренд
    cars_without_rentals = all_cars.exclude(id__in=cars_with_rentals)

    shifts_dict = {}
    for shift in shifts:
        key = (shift.rental_id, shift.date)
        shifts_dict.setdefault(key, []).append(shift)

    # Начинаем заполнять таблицу данными
    current_row = 2
    for rental in rentals:
        total_hours = 0
        total_amount = 0
        # Заполняем данные аренды
        ws.cell(row=current_row, column=1, value=rental.id)
        ws.cell(row=current_row, column=2, value=rental.car.name)
        ws.cell(row=current_row, column=3, value=rental.build_object.name if rental.build_object else '')
        ws.cell(row=current_row, column=4, value=rental.tariff if rental.tariff else '')
        ws.cell(row=current_row, column=5, value=rental.extra_tariff if rental.extra_tariff else '')

        # Перебираем дни месяца
        current_date = start_date
        while current_date <= end_date:
            hours_worked = 0
            day_amount = 0
            if rental.start_date <= current_date.date() <= rental.end_date:
                temp_shift = shifts_dict.get((rental.id, current_date.date()))
                if temp_shift:
                    for a in temp_shift:
                        start_time = a.start_time
                        end_time = a.end_time
                        hours = (end_time.hour - start_time.hour) + (end_time.minute - start_time.minute) / 60
                        day_amount += hours * rental.extra_tariff if a.is_additional_mode else hours * rental.tariff
                        hours_worked += hours
                    if any(_.is_additional_mode for _ in temp_shift):
                        cell = ws.cell(row=current_row, column=current_date.day + 5)
                        cell.value = hours_worked
                        cell.fill = blue_fill  # Синий
                    else:
                        ws.cell(row=current_row, column=current_date.day + 5, value=hours_worked)
                        cell = ws.cell(row=current_row, column=current_date.day + 5)
                        cell.fill = green_fill  # Зеленый
                else:
                    # Если нет смены, то помечаем красным и ставим 0
                    cell = ws.cell(row=current_row, column=current_date.day + 5)
                    cell.fill = red_fill  # Красный
                    cell.value = 0
            # else:
            #     ws.cell(row=current_row, column=current_date.day + 5, value='')

            current_date += timedelta(days=1)
            total_hours += hours_worked
            total_amount += day_amount
        ws.cell(row=current_row, column=num_days + 6, value=total_hours)
        ws.cell(row=current_row, column=num_days + 7, value=total_amount)

        current_row += 1

    # Примените стиль к строке с меню
    menu_style = NamedStyle(name="menu_style")
    menu_style.font = menu_font
    menu_style.fill = menu_fill
    menu_style.alignment = Alignment(horizontal="center", vertical="center")

    menu_row = ws[1]
    for cell in menu_row:
        cell.style = menu_style
        cell.border = medium_border  # Средняя граница для меню

    # Примените стиль к данным ячеек
    for row in ws.iter_rows(min_row=2, max_row=current_row - 1):
        for cell in row:
            cell.font = cell_font
            cell.border = thin_border  # Тонкие границы для данных ячеек

    # Примените стиль к столбцам с итоговыми часами и ценой
    for col in range(num_days + 6, num_days + 8):
        column_letter = get_column_letter(col)
        for cell in ws[column_letter]:
            cell.font = total_font
            cell.border = thin_border  # Тонкие границы для итоговых ячеек

    for car in cars_without_rentals:
        ws.cell(row=current_row, column=2, value=car.name)
        current_row += 1

    for column in ws.columns:
        column_name = get_column_letter(column[0].column)  # Получаем имя колонки (например, A, B, C)
        max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
        adjusted_width = (max_length + 2) * 1.5  # Расширяем колонку с небольшим запасом
        ws.column_dimensions[column_name].width = adjusted_width

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=rental_table_{selected_date.strftime("%Y-%m")}.xlsx'
    wb.save(response)

    return response


class CalendarService:
    @staticmethod
    def get_booked_dates(selected_month):
        redis_conn = get_redis_connection()
        cache_key = f"rental_calendar_{selected_month}"
        calendar = redis_conn.get(cache_key)

        if calendar:
            return json.loads(calendar)
        else:
            return CalendarService.calculate_and_cache_booked_dates(selected_month)

    @staticmethod
    def calculate_and_cache_booked_dates(selected_month):
        selected_date = timezone.datetime.strptime(selected_month, '%Y-%m').date()
        first_day_of_month = selected_date.replace(day=1)
        last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        rentals = Rental.objects.filter(
            end_date__gte=first_day_of_month,
            start_date__lte=last_day_of_month
        )

        booked_dates = {}
        cars = Car.objects.all()
        for car in cars:
            booked_dates[car.name] = []

        for rental in rentals:
            car_name = rental.car.name
            rental_data = {
                rental.id: [rental.client.name, [date.strftime('%m-%d') for date in
                                                 (rental.start_date + timedelta(days=i) for i in
                                                  range((rental.end_date - rental.start_date).days + 1))]]
            }
            booked_dates[car_name].append(rental_data)

        redis_conn = get_redis_connection()
        cache_key = f"rental_calendar_{selected_month}"
        redis_conn.set(cache_key, json.dumps(booked_dates), ex=3600)

        return booked_dates
